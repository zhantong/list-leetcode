import urllib.request
import urllib.parse
import http.cookiejar
import re
import json
import csv


class LeetCode:
    def __init__(self):
        self.base_url = 'https://leetcode.com/'
        cj = http.cookiejar.CookieJar()
        self.opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
        self.opener.addheaders = [
            ('Host', 'leetcode.com'),
            ('User-Agent',
             'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36'),
            ('Referer', 'https://leetcode.com/accounts/login/')
        ]

    def login_from_config(self):
        with open('config.json', encoding='utf-8') as f:
            config = json.loads(f.read())
        return self.login(config['UserName'], config['Password'])

    def login(self, user_name, password):
        with self.opener.open(self.base_url + 'accounts/login/') as f:
            content = f.read().decode('utf-8')
        token = re.findall("name='csrfmiddlewaretoken'\svalue='(.*?)'", content)[0]
        post_data = {
            'csrfmiddlewaretoken': token,
            'login': user_name,
            'password': password
        }
        post_data = urllib.parse.urlencode(post_data)
        with self.opener.open(self.base_url + 'accounts/login/', data=post_data.encode()) as f:
            if f.read().decode().find('Successfully signed in') != -1:
                print('Successfully signed in')
            else:
                print('Failed sign in')

    def get_problem_list(self):
        with self.opener.open(self.base_url + 'api/problems/algorithms/') as f:
            content = f.read().decode('utf-8')
        content = json.loads(content)
        result = []
        for problem in content['stat_status_pairs']:
            result.append({
                'id': problem['stat']['question_id'],
                'title': problem['stat']['question__title'],
                'slug': problem['stat']['question__title_slug'],
                'difficulty': problem['difficulty']['level'],
                'total_submitted': problem['stat']['total_submitted'],
                'total_acs': problem['stat']['total_acs'],
                'acceptance': problem['stat']['total_acs'] / problem['stat']['total_submitted'],
                'paid_only': problem['paid_only'],
                'status': True if problem['status'] else False,
            })
        result.sort(key=lambda x: x['id'])
        return result

    def to_locale(self, problem_list, language_dict):
        problem_list = [{language_dict[key]: value for (key, value) in problem.items()} for problem in problem_list]

        for problem in problem_list:
            problem[language_dict['difficulty']] = language_dict['level'][problem[language_dict['difficulty']]]
            problem[language_dict['paid_only']] = language_dict['bool'][problem[language_dict['paid_only']]]
            problem[language_dict['status']] = language_dict['bool'][problem[language_dict['status']]]
        return problem_list

    def save_problem_list_as_csv(self, problem_list, file_name):
        with open(file_name, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=problem_list[0].keys())
            writer.writeheader()
            writer.writerows(problem_list)

    def save_problem_list_as_excel(self, problem_list, file_name, language_dict):
        from openpyxl import Workbook
        from openpyxl.styles import NamedStyle
        from openpyxl.formatting.rule import CellIsRule, DataBarRule
        from openpyxl.styles import PatternFill

        def format_cell_style(ws, language_dict):
            style_int = NamedStyle('int')
            style_int.number_format = '0'
            style_str = NamedStyle('str')
            style_str.number_format = '@'
            style_pcnt = NamedStyle('pcnt')
            style_pcnt.number_format = '0.0%'
            for cell in ws[column_index[language_dict['id']]][1:]:
                cell.style = style_int
            for cell in ws[column_index[language_dict['total_submitted']]][1:]:
                cell.style = style_int
            for cell in ws[column_index[language_dict['total_acs']]][1:]:
                cell.style = style_int
            for cell in ws[column_index[language_dict['title']]][1:]:
                cell.style = style_str
            for cell in ws[column_index[language_dict['slug']]][1:]:
                cell.style = style_str
            for cell in ws[column_index[language_dict['difficulty']]][1:]:
                cell.style = style_str
            for cell in ws[column_index[language_dict['paid_only']]][1:]:
                cell.style = style_str
            for cell in ws[column_index[language_dict['status']]][1:]:
                cell.style = style_str
            for cell in ws[column_index[language_dict['acceptance']]][1:]:
                cell.style = style_pcnt

        def conditional_formatting(ws, language_dict):
            red_color = 'ffc7ce'
            green_color = 'c2efcf'
            yellow_color = 'ffeba2'

            red_fill = PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
            green_fill = PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
            yellow_fill = PatternFill(start_color=yellow_color, end_color=yellow_color, fill_type='solid')

            ws.conditional_formatting.add(self.get_entire_column(column_index[language_dict['difficulty']]),
                                          CellIsRule(operator='equal', formula=['"' + language_dict['level'][1] + '"'],
                                                     stopIfTrue=False, fill=green_fill))
            ws.conditional_formatting.add(self.get_entire_column(column_index[language_dict['difficulty']]),
                                          CellIsRule(operator='equal', formula=['"' + language_dict['level'][2] + '"'],
                                                     stopIfTrue=False, fill=yellow_fill))
            ws.conditional_formatting.add(self.get_entire_column(column_index[language_dict['difficulty']]),
                                          CellIsRule(operator='equal', formula=['"' + language_dict['level'][3] + '"'],
                                                     stopIfTrue=False, fill=red_fill))

            ws.conditional_formatting.add(self.get_entire_column(column_index[language_dict['paid_only']]),
                                          CellIsRule(operator='equal',
                                                     formula=['"' + language_dict['bool'][False] + '"'],
                                                     stopIfTrue=False, fill=green_fill))
            ws.conditional_formatting.add(self.get_entire_column(column_index[language_dict['paid_only']]),
                                          CellIsRule(operator='equal',
                                                     formula=['"' + language_dict['bool'][True] + '"'],
                                                     stopIfTrue=False, fill=red_fill))

            ws.conditional_formatting.add(self.get_entire_column(column_index[language_dict['status']]),
                                          CellIsRule(operator='equal',
                                                     formula=['"' + language_dict['bool'][False] + '"'],
                                                     stopIfTrue=False, fill=red_fill))
            ws.conditional_formatting.add(self.get_entire_column(column_index[language_dict['status']]),
                                          CellIsRule(operator='equal',
                                                     formula=['"' + language_dict['bool'][True] + '"'],
                                                     stopIfTrue=False, fill=green_fill))

            ws.conditional_formatting.add(self.get_entire_column(column_index[language_dict['acceptance']]),
                                          DataBarRule(start_type='percentile', start_value=0, end_type='percentile',
                                                      end_value=100, color="FF638EC6", showValue='None'))

        wb = Workbook()
        ws = wb.active
        ws.append(tuple(problem_list[0].keys()))
        column_index = {item.value: item.column for item in ws[1]}
        rows = [{column_index[key]: value for (key, value) in problem.items()} for problem in problem_list]
        for row in rows:
            ws.append(row)
        format_cell_style(ws, language_dict)
        conditional_formatting(ws, language_dict)
        wb.save(file_name)

    def get_entire_column(self, index):
        return index + '1:' + index + '1048576'

    def load_data(self, file_path):
        with open(file_path, encoding='utf-8') as f:
            return json.loads(f.read())

    def get_language_dict(self, language):
        language_dict = None
        if language == 'Chinese':
            language_dict = {
                'id': '题号',
                'title': '标题',
                'slug': '链接',
                'difficulty': '难度',
                'total_submitted': '总提交数',
                'total_acs': '总通过数',
                'acceptance': '通过率',
                'paid_only': '付费',
                'status': '已解决',
                'level': {
                    1: '简单',
                    2: '中等',
                    3: '难'
                },
                'bool': {
                    True: '是',
                    False: '否'
                }
            }
        elif language == 'English':
            language_dict = {
                'id': '#',
                'title': 'Title',
                'slug': 'Link',
                'difficulty': 'Difficulty',
                'total_submitted': 'Total Submitted',
                'total_acs': 'Total Accepted',
                'acceptance': 'Acceptance',
                'paid_only': 'Paid Only',
                'status': 'Solved',
                'level': {
                    1: 'Easy',
                    2: 'Medium',
                    3: 'Hard'
                },
                'bool': {
                    True: 'Yes',
                    False: 'No'
                }
            }
        return language_dict
